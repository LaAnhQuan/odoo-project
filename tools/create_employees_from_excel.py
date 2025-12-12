"""
Script tạo nhân viên từ file Excel vào Odoo
Chạy trong môi trường Odoo shell
"""

import openpyxl

# Đường dẫn file Excel của bạn
excel_file = r"D:\path\to\your\T7-BTGD 1.xlsx"

# Load workbook
wb = openpyxl.load_workbook(excel_file, data_only=True)
ws = wb.active

# Cấu hình đọc
data_start_row = 9  # Bắt đầu từ dòng 9
col_mans = 2        # Cột B
col_name = 3        # Cột C

# Lấy model
Employee = env['hr.employee']
company = env.company

created = []
skipped = []

print("Bắt đầu tạo nhân viên từ Excel...")

for r in range(data_start_row, min(data_start_row + 100, ws.max_row + 1)):
    mans = ws.cell(r, col_mans).value
    name = ws.cell(r, col_name).value
    
    if mans:
        mans = str(mans).strip()
    if name:
        name = str(name).strip()
    
    # Bỏ qua dòng không có MANS
    if not mans:
        continue
    
    # Bỏ qua dòng footer
    if name and any(kw in name.upper() for kw in ['CỘNG THÁNG', 'LẬP BẢNG', 'KHÔNG XOÁ']):
        skipped.append(f"Dòng {r}: Footer '{name}'")
        continue
    
    # Kiểm tra đã tồn tại chưa
    existing = Employee.search([('mans', '=', mans), ('company_id', '=', company.id)], limit=1)
    if existing:
        skipped.append(f"Dòng {r}: MANS '{mans}' đã tồn tại - {existing.name}")
        continue
    
    # Tạo tên mặc định nếu không có
    if not name:
        name = f"Nhân viên {mans}"
    
    # Tạo nhân viên
    try:
        emp = Employee.create({
            'name': name,
            'mans': mans,
            'company_id': company.id,
        })
        created.append(f"✅ Dòng {r}: {mans} - {name}")
        print(f"✅ Tạo: {mans} - {name}")
    except Exception as e:
        skipped.append(f"❌ Dòng {r}: Lỗi - {str(e)}")

print("\n" + "="*60)
print(f"KẾT QUẢ:")
print(f"✅ Tạo thành công: {len(created)} nhân viên")
print(f"⚠️  Bỏ qua: {len(skipped)} dòng")

if created:
    print("\n--- Nhân viên đã tạo ---")
    for c in created:
        print(c)

if skipped:
    print("\n--- Dòng bỏ qua ---")
    for s in skipped[:10]:  # Chỉ hiển thị 10 dòng đầu
        print(s)
    if len(skipped) > 10:
        print(f"... và {len(skipped) - 10} dòng khác")
