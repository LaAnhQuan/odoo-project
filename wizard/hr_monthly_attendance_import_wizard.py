from odoo import api, fields, models, _
from odoo.exceptions import UserError
from datetime import date
import base64
import calendar
from odoo.tools import misc

try:
    import openpyxl
except Exception:
    openpyxl = None


def _norm(v):
    return (str(v).strip() if v is not None else "")


def _map_cell_to_code(cell_value):
    v = _norm(cell_value).upper().replace(" ", "")
    # trống = công
    if v == "":
        return "W"

    mapping = {
        "P": "P",
        "P/2": "P2",
        "P2": "P2",
        "KO": "KO",
        "KO/2": "KO2",
        "KO2": "KO2",
        "OFF": "OFF",
        "X": "W",
        "C": "W",
        # nhiều file ghi Ko/2 kiểu chữ thường
        "KO/2": "KO2",
        "KO/2": "KO2",
    }
    return mapping.get(v, "W")


class HrMonthlyAttendanceImportWizard(models.TransientModel):
    _name = "hr.monthly.attendance.import.wizard"
    _description = "Import Monthly Attendance Matrix Excel"

    sheet_id = fields.Many2one("hr.monthly.attendance", required=True, ondelete="cascade")
    file_data = fields.Binary(string="File Excel", required=True)
    file_name = fields.Char(string="Tên file")

    # Layout cố định theo file khách
    header_row = fields.Integer(default=8)
    data_start_row = fields.Integer(default=10)  # Dòng 9 là thứ, dòng 10 mới là data
    col_mans = fields.Integer(default=2)      # B
    col_name = fields.Integer(default=3)      # C
    col_day_01 = fields.Integer(default=4)    # D

    def action_import(self):
        self.ensure_one()
        if openpyxl is None:
            raise UserError(_("Thiếu thư viện openpyxl. Cài: pip install openpyxl"))

        sheet = self.sheet_id
        if not sheet.month or not sheet.year:
            raise UserError(_("Bảng tháng chưa có tháng/năm."))

        month_int = int(sheet.month)
        last_day = calendar.monthrange(sheet.year, month_int)[1]

        # load workbook từ binary
        import io
        content = base64.b64decode(self.file_data)
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        Employee = self.env["hr.employee"].sudo()
        Daily = self.env["hr.daily.attendance"].sudo()

        created = 0
        updated = 0
        not_found = []
        skipped_rows = []
        imported_employees = 0

        # Từ khóa để nhận diện dòng footer cần bỏ qua
        footer_keywords = ['CỘNG THÁNG', 'CÔNG THÁNG', 'LẬP BẢNG', 'KHÔNG XOÁ', 'TỔNG', 'GHI CHÚ']

        # đọc từng dòng nhân viên
        for r in range(self.data_start_row, min(self.data_start_row + 200, ws.max_row + 1)):
            mans = _norm(ws.cell(r, self.col_mans).value)
            emp_name = _norm(ws.cell(r, self.col_name).value)

            # Bỏ qua dòng trống hoàn toàn
            if not mans and not emp_name:
                continue

            # Bỏ qua dòng footer (Cộng tháng, Lập bảng...)
            emp_name_upper = emp_name.upper()
            if any(keyword in emp_name_upper for keyword in footer_keywords):
                skipped_rows.append(f"Dòng {r}: Footer - '{emp_name}'")
                continue

            # Bỏ qua dòng có MANS = số 0 hoặc chỉ có số (không phải mã hợp lệ)
            if mans and mans.isdigit() and int(mans) == 0:
                skipped_rows.append(f"Dòng {r}: MANS không hợp lệ '{mans}'")
                continue

            # tìm employee theo mans (ưu tiên)
            emp = False
            if mans:
                emp = Employee.search([
                    ("mans", "=", mans),
                    ("company_id", "=", sheet.company_id.id),
                ], limit=1)

            # fallback theo tên (cẩn thận trùng)
            if not emp and emp_name:
                emps = Employee.search([
                    ("name", "ilike", emp_name),
                    ("company_id", "=", sheet.company_id.id),
                ])
                if len(emps) == 1:
                    emp = emps
                elif len(emps) > 1:
                    not_found.append(f"Dòng {r}: Trùng tên '{emp_name}' - Cần MANS để phân biệt")
                    continue

            if not emp:
                not_found.append(f"Dòng {r}: MANS='{mans}', Họ tên='{emp_name}'")
                continue

            # Đọc dữ liệu chấm công 01..last_day
            imported_employees += 1

            # đọc 01..last_day
            for d in range(1, last_day + 1):
                c = self.col_day_01 + (d - 1)
                cell_val = ws.cell(r, c).value
                code = _map_cell_to_code(cell_val)

                att_date = date(sheet.year, month_int, d)

                # upsert theo unique(employee_id, date)
                rec = Daily.search([("employee_id", "=", emp.id), ("date", "=", att_date)], limit=1)
                vals = {
                    "employee_id": emp.id,
                    "date": att_date,
                    "attendance_code": code,
                }
                if rec:
                    rec.write(vals)
                    updated += 1
                else:
                    Daily.create(vals)
                    created += 1

        # Tạo thông báo kết quả
        message_parts = []
        message_parts.append(f"✅ Import thành công {imported_employees} nhân viên")
        message_parts.append(f"📝 Tạo mới: {created} | Cập nhật: {updated}")
        
        if skipped_rows:
            message_parts.append(f"\n⚠️ Bỏ qua {len(skipped_rows)} dòng (footer/không hợp lệ)")
        
        if not_found:
            # Chỉ hiển thị 10 dòng đầu để không quá dài
            preview_not_found = not_found[:10]
            message_parts.append(f"\n❌ Không tìm thấy {len(not_found)} nhân viên:")
            message_parts.append("\n".join([f"  • {nf}" for nf in preview_not_found]))
            if len(not_found) > 10:
                message_parts.append(f"  ... và {len(not_found) - 10} nhân viên khác")
            message_parts.append("\n💡 Cần tạo nhân viên với MANS tương ứng trong HR → Employees")

        full_message = "\n".join(message_parts)

        # Nếu không import được nhân viên nào, báo lỗi
        if imported_employees == 0:
            raise UserError(_("Không import được nhân viên nào!\n\n%s") % full_message)

        # Thành công một phần hoặc toàn bộ
        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Hoàn thành Import"),
                "message": full_message,
                "sticky": True,
                "type": "warning" if not_found else "success",
            },
        }
