from odoo import api, fields, models, _
from odoo.exceptions import UserError
from odoo.modules.module import get_module_resource

from datetime import date
import calendar
import io
import base64

try:
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill
except Exception:
    openpyxl = None


# map code -> value (đồng bộ với hr.daily.attendance)
CODE_TO_VALUE = {"W": 1.0, "P": 1.0, "P2": 0.5, "KO": 0.0, "KO2": 0.5, "OFF": 0.0}


class HrMonthlyAttendance(models.Model):
    _name = "hr.monthly.attendance"
    _description = "Monthly Attendance Sheet"
    _inherit = ["mail.thread", "mail.activity.mixin"]

    name = fields.Char(string="Tên bảng", compute="_compute_name", store=True)

    month = fields.Selection(
        [(str(i), f"Tháng {i}") for i in range(1, 13)],
        string="Tháng",
        required=True,
        tracking=True,
    )
    year = fields.Integer(
        string="Năm",
        required=True,
        default=lambda self: fields.Date.today().year,
        tracking=True,
    )

    company_id = fields.Many2one(
        "res.company",
        string="Công ty",
        default=lambda self: self.env.company,
        required=True,
    )
    department_id = fields.Many2one(
        "hr.department",
        string="Phòng ban",
        help="Nếu để trống sẽ tính cho tất cả nhân viên của công ty.",
    )

    state = fields.Selection(
        [
            ("draft", "Nháp"),
            ("computed", "Đã tính từ chấm công"),
            ("confirmed", "Đã xác nhận"),
            ("done", "Đã chuyển sang lương"),
        ],
        string="Trạng thái",
        default="draft",
        tracking=True,
    )

    line_ids = fields.One2many("hr.monthly.attendance.line", "sheet_id", string="Chi tiết chấm công")

    date_from = fields.Date(string="Từ ngày", compute="_compute_date_range", store=True)
    date_to = fields.Date(string="Đến ngày", compute="_compute_date_range", store=True)

    note = fields.Text(string="Ghi chú")

    # file export
    export_file = fields.Binary(string="Export file", readonly=True)
    export_filename = fields.Char(string="Tên file export", readonly=True)

    # ------------------------------------------------------------
    # COMPUTE
    # ------------------------------------------------------------
    @api.depends("month", "year")
    def _compute_name(self):
        for rec in self:
            if rec.month and rec.year:
                rec.name = f"Bảng công tháng {rec.month}/{rec.year}"
            else:
                rec.name = "Bảng công"

    @api.depends("month", "year")
    def _compute_date_range(self):
        for rec in self:
            if rec.month and rec.year:
                m = int(rec.month)
                last_day = calendar.monthrange(rec.year, m)[1]
                rec.date_from = date(rec.year, m, 1)
                rec.date_to = date(rec.year, m, last_day)
            else:
                rec.date_from = False
                rec.date_to = False

    # ------------------------------------------------------------
    # ACTION: Compute from hr.daily.attendance
    # ------------------------------------------------------------
    def action_compute_from_attendance(self):
        """
        Tính bảng tháng từ hr.daily.attendance:
        - worked_days: tổng workday_value (W=1, P=1, P2=0.5, KO=0, KO2=0.5, OFF=0)
        - paid_leave_days: tổng ngày phép (P=1, P2=0.5)
        - unpaid_leave_days: tổng KO/KO2/OFF theo value tương ứng (KO=1? -> ở bạn đang quy KO=0.0 nên để riêng)
          => Thực tế khách hay muốn: nghỉ không lương = (KO=1.0, KO2=0.5). Nếu muốn vậy thì đổi map.
        - overtime_hours: sum overtime_hours
        """
        Daily = self.env["hr.daily.attendance"]

        for sheet in self:
            if not sheet.date_from or not sheet.date_to:
                continue

            # clear cũ
            if sheet.line_ids:
                sheet.line_ids.mapped("daily_ids").write({"monthly_sheet_id": False, "monthly_line_id": False})
                sheet.line_ids.unlink()

            domain = [
                ("date", ">=", sheet.date_from),
                ("date", "<=", sheet.date_to),
                ("company_id", "=", sheet.company_id.id),
            ]
            if sheet.department_id:
                domain.append(("department_id", "=", sheet.department_id.id))

            dailies = Daily.search(domain)

            # group by employee
            data = {}
            for d in dailies:
                emp = d.employee_id
                if not emp:
                    continue
                data.setdefault(emp.id, {
                    "worked": 0.0,
                    "paid_leave": 0.0,
                    "unpaid_leave": 0.0,
                    "ot": 0.0,
                    "daily_ids": self.env["hr.daily.attendance"],
                })

                code = d.attendance_code or "W"
                
                # Kiểm tra ngày trong tuần
                is_saturday = d.date.weekday() == 5
                is_sunday = d.date.weekday() == 6
                
                # Tính giá trị ngày công dựa vào code và ngày trong tuần
                if code in ("W", "X"):
                    # Ngày công thực tế (đi làm)
                    if is_sunday:
                        val = 0.0  # Chủ nhật không tính công
                    elif is_saturday:
                        val = 0.5  # Thứ 7 chỉ tính nửa công
                    else:
                        val = 1.0  # Ngày thường tính đủ
                    data[emp.id]["worked"] += val
                    
                elif code in ("P", "P2"):
                    # Ngày nghỉ hưởng lương (Phép)
                    val = CODE_TO_VALUE.get(code, 1.0)
                    data[emp.id]["worked"] += val
                    data[emp.id]["paid_leave"] += val
                    
                elif code in ("KO", "KO2", "OFF"):
                    # Ngày nghỉ không lương
                    val = CODE_TO_VALUE.get(code, 0.0)
                    if code in ("KO", "KO2"):
                        # KO = 1 ngày không lương, KO2 = 0.5 ngày không lương
                        unpaid_val = 1.0 if code == "KO" else 0.5
                        data[emp.id]["unpaid_leave"] += unpaid_val
                else:
                    # Code khác (nếu có) - sử dụng giá trị mặc định
                    val = d.workday_value if d.workday_value is not None else CODE_TO_VALUE.get(code, 1.0)
                    data[emp.id]["worked"] += float(val)

                data[emp.id]["ot"] += float(d.overtime_hours or 0.0)
                data[emp.id]["daily_ids"] |= d

            # tạo line
            for emp_id, v in data.items():
                line = self.env["hr.monthly.attendance.line"].create({
                    "sheet_id": sheet.id,
                    "employee_id": emp_id,
                    "worked_days": v["worked"],
                    "paid_leave_days": v["paid_leave"],
                    "unpaid_leave_days": v["unpaid_leave"],
                    "overtime_hours": v["ot"],
                })

                v["daily_ids"].write({"monthly_sheet_id": sheet.id, "monthly_line_id": line.id})

            sheet.state = "computed"

    def action_confirm(self):
        for sheet in self:
            sheet.state = "confirmed"

    def action_done(self):
        for sheet in self:
            sheet.state = "done"

    def action_set_to_draft(self):
        for sheet in self:
            sheet.state = "draft"

    # ------------------------------------------------------------
    # ACTION: Open Grid View
    # ------------------------------------------------------------
    def action_open_grid_view(self):
        """Mở grid view với dữ liệu của sheet này"""
        self.ensure_one()
        
        # Đồng bộ dữ liệu sang grid view
        Grid = self.env["hr.monthly.attendance.grid"]
        Grid.sync_all_from_monthly_sheet(self.id)
        
        return {
            "type": "ir.actions.client",
            "tag": "monthly_attendance_grid",
            "name": f"Grid - {self.name}",
            "context": {
                "default_month": self.month,
                "default_year": self.year,
                "default_department_id": self.department_id.id if self.department_id else False,
            },
        }

    # ------------------------------------------------------------
    # ACTION: Open Import Wizard (Excel matrix)
    # ------------------------------------------------------------
    def action_open_import_matrix(self):
        self.ensure_one()
        return {
            "type": "ir.actions.act_window",
            "name": _("Import Excel (Matrix)"),
            "res_model": "hr.monthly.attendance.import.wizard",
            "view_mode": "form",
            "target": "new",
            "context": {"default_sheet_id": self.id},
        }

    # ------------------------------------------------------------
    # ACTION: Export XLSX using template (GIỮ NGUYÊN FORMAT)
    # ------------------------------------------------------------
    def action_export_xlsx_matrix(self):
        """
        Export y chang file khách hàng bằng cách:
        - load template xlsx trong module
        - đổ dữ liệu vào đúng vùng:
            header row = 8
            data row start = 9
            MANS col = B (2)
            Họ tên col = C (3)
            Day 01 col = D (4)
        - giữ nguyên merge/màu/logo/CN xám...
        """
        self.ensure_one()
        if openpyxl is None:
            raise UserError(_("Thiếu thư viện openpyxl. Cài: pip install openpyxl"))

        sheet = self
        if not sheet.month or not sheet.year:
            raise UserError(_("Bảng tháng chưa có tháng/năm."))

        # 1) load template
        # Bạn đặt template vào: <module>/static/src/xlsx/monthly_template.xlsx
        template_path = get_module_resource(
            "mo_hr_monthly_attendance",  # <-- đổi đúng tên technical module của bạn
            "static", "src", "xlsx", "monthly_template.xlsx"
        )
        if not template_path:
            raise UserError(_("Không tìm thấy template monthly_template.xlsx trong module."))

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        header_row = 8
        weekday_row = 9   # Dòng thứ trong tuần
        start_row = 10    # Dòng data nhân viên bắt đầu từ 10
        mans_col = 2   # B
        name_col = 3   # C
        day1_col = 4   # D

        month_int = int(sheet.month)
        last_day = calendar.monthrange(sheet.year, month_int)[1]

        # Tạo dòng thứ trong tuần (row 9)
        weekdays = ['CN', 'T2', 'T3', 'T4', 'T5', 'T6', 'T7']
        for d in range(1, last_day + 1):
            day_date = date(sheet.year, month_int, d)
            weekday_name = weekdays[day_date.weekday() + 1] if day_date.weekday() < 6 else 'CN'
            cell = ws.cell(weekday_row, day1_col + (d - 1))
            cell.value = weekday_name
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # Tô xám Chủ nhật
            if weekday_name == 'CN':
                cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

        # 2) đổ dữ liệu daily -> matrix theo mans
        Daily = self.env["hr.daily.attendance"].sudo()
        domain = [
            ("date", ">=", sheet.date_from),
            ("date", "<=", sheet.date_to),
            ("company_id", "=", sheet.company_id.id),
        ]
        if sheet.department_id:
            domain.append(("department_id", "=", sheet.department_id.id))

        dailies = Daily.search(domain)
        by_emp = {}
        for d in dailies:
            emp = d.employee_id
            if not emp:
                continue
            key = emp.id
            by_emp.setdefault(key, {"emp": emp, "days": {}})
            by_emp[key]["days"][d.date.day] = (d.attendance_code or "W")

        # 3) clear vùng cũ (trong template có thể còn dữ liệu mẫu)
        max_rows_to_clear = 300
        for r in range(start_row, start_row + max_rows_to_clear):
            # nếu muốn dừng khi hết bảng mẫu, bạn có thể break khi cột C trống
            # nhưng template có thể trống sẵn, nên mình clear toàn vùng vừa đủ.
            for d in range(1, 32):
                ws.cell(r, day1_col + (d - 1)).value = None
            # cột tổng bên phải nếu template có (Công/Phép/Ko...) bạn tuỳ add sau

        # 4) write rows
        employees = list(by_emp.values())

        # sort theo mans rồi tên
        employees.sort(key=lambda x: (x["emp"].mans or "", x["emp"].name or ""))

        r = start_row
        for idx, item in enumerate(employees, start=1):
            emp = item["emp"]
            ws.cell(r, 1).value = idx         # STT ở cột A (nếu template có)
            ws.cell(r, mans_col).value = emp.mans or ""
            ws.cell(r, name_col).value = emp.name or ""

            for d in range(1, last_day + 1):
                code = item["days"].get(d, "")
                # quy ước template: trống = công, P/P2/KO/KO2/OFF đúng text
                if code == "W":
                    ws.cell(r, day1_col + (d - 1)).value = ""
                elif code == "P2":
                    ws.cell(r, day1_col + (d - 1)).value = "P/2"
                elif code == "KO2":
                    ws.cell(r, day1_col + (d - 1)).value = "Ko/2"
                else:
                    ws.cell(r, day1_col + (d - 1)).value = code
            r += 1

        # 5) output
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        filename = f"Bang_cham_cong_{sheet.month}_{sheet.year}.xlsx"
        sheet.write({
            "export_file": base64.b64encode(out.read()),
            "export_filename": filename,
        })

        return {
            "type": "ir.actions.act_url",
            "url": "/web/content/?model=hr.monthly.attendance"
                   f"&id={sheet.id}&field=export_file&filename_field=export_filename"
                   "&download=true",
            "target": "self",
        }
    
    # def action_open_import_matrix(self):
    #     self.ensure_one()
    #     return {
    #         "type": "ir.actions.act_window",
    #         "name": "Import Excel (Ma trận)",
    #         "res_model": "hr.monthly.attendance.import.wizard",
    #         "view_mode": "form",
    #         "target": "new",
    #         "context": {
    #             "default_sheet_id": self.id,
    #         },
    #     }



class HrMonthlyAttendanceLine(models.Model):
    _name = "hr.monthly.attendance.line"
    _description = "Monthly Attendance Line"

    sheet_id = fields.Many2one(
        "hr.monthly.attendance",
        string="Bảng chấm công tháng",
        ondelete="cascade",
        required=True,
    )

    employee_id = fields.Many2one("hr.employee", string="Nhân viên", required=True)

    mans = fields.Char(string="MANS", related="employee_id.mans", store=True, index=True, readonly=True)

    daily_ids = fields.One2many("hr.daily.attendance", "monthly_line_id", string="Chi tiết chấm công ngày")

    worked_days = fields.Float(string="Ngày công thực tế")
    paid_leave_days = fields.Float(string="Ngày nghỉ hưởng lương")
    unpaid_leave_days = fields.Float(string="Ngày nghỉ không lương")
    overtime_hours = fields.Float(string="Giờ làm thêm")

    total_paid_days = fields.Float(
        string="Tổng ngày công tính lương",
        compute="_compute_total_paid_days",
        store=True,
    )

    note = fields.Char(string="Ghi chú")

    @api.depends("worked_days", "paid_leave_days")
    def _compute_total_paid_days(self):
        for line in self:
            line.total_paid_days = (line.worked_days or 0.0) + (line.paid_leave_days or 0.0)
